# -*- coding: utf-8 -*-
"""Tests d'intégration : classeurs Excel générés à la volée, chaîne complète."""

import os
import shutil
from datetime import datetime

import pytest
from openpyxl import Workbook

from excel_to_quadra.cli import main as cli_main
from excel_to_quadra.config import (Composante, Configuration, Source, SourcePaie)
from excel_to_quadra.moteur import (EnteteInvalide, archiver_entree,
                                     controler_equilibre, ecrire_fichiers,
                                     generer_ecritures, generer_ecritures_paie,
                                     nettoyer_sortie, source_correspond)

CRLF = b"\r\n"


@pytest.fixture
def environnement(tmp_path):
    """Crée deux classeurs sources (simple + paie) et la configuration associée."""
    entree = tmp_path / "entree"
    sortie = tmp_path / "sortie"
    entree.mkdir()

    # --- Classeur « une ligne = un établissement » (type OETH) ---
    wb = Workbook()
    ws = wb.active
    ws.title = "CRE"
    ws["A1"] = "Provision"                       # en-têtes hors zone de données
    donnees = [("7704 - CRECHE A", 100.0), ("7705 - CRECHE B", 0.0),
               ("7799 - SANS CENTRE", 50.0), ("Total", 150.0)]
    for i, (code, montant) in enumerate(donnees, start=3):
        ws.cell(i, 1, code)
        ws.cell(i, 3, montant)
    wb.save(entree / "provision.xlsx")

    # --- Classeur de paie détaillé par salarié (type PRECA) ---
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Feuil1"
    paie = [  # (centre de coût, prime) — 2 salariés même centre + 1 négatif + 1 inconnu
        ("770401", 60.0), ("770401", 40.0), ("770501", -30.0), ("999999", 10.0)]
    for i, (centre, prime) in enumerate(paie, start=2):
        ws2.cell(i, 4, centre)
        ws2.cell(i, 14, prime)                    # colonne N
    wb2.save(entree / "preca.xlsx")

    cfg = Configuration(
        dossier_entree=str(entree),
        dossier_sortie=str(sortie),
        analytique={"704": "770401", "705": "770501"},
        centre_vers_dossier={"770401": "704", "770501": "705"},
        sources=[Source(
            fichier="provision.xlsx", feuille="CRE", ligne_debut=3,
            col_dossier="A", col_montant="C", extraire_code=True,
            compte_credit="43785000", compte_debit="63388100",
            libelle="OETH TEST", journal="OS", date_ecriture="310526",
            contre_passation="010626")],
        sources_paie=[SourcePaie(
            fichier="preca.xlsx", feuille="Feuil1", ligne_debut=2, col_centre="D",
            journal="OS", date_ecriture="310526", contre_passation=None,
            composantes=[Composante(col="N", compte_debit="64133820",
                                    compte_credit="42822000", libelle="PRIME TEST")])],
    )
    return cfg


class TestChaineComplete:
    def test_lignes_filtrees_et_dossiers_produits(self, environnement):
        par_dossier, sans_centre = generer_ecritures(environnement.sources, environnement)
        assert "704" in par_dossier                 # 100 € comptabilisés
        assert "705" not in par_dossier             # montant nul écarté
        assert "799" in par_dossier                 # écriture produite même sans centre
        assert ("799", "OETH TEST", "provision.xlsx") in sans_centre  # ... mais signalée

    def test_agregation_paie_et_routage(self, environnement):
        par_paie, inconnus, attente = generer_ecritures_paie(
            environnement.sources_paie, environnement)
        # 60 + 40 agrégés sur le centre 770401 -> dossier 704
        debit = next(l for l in par_paie["704"] if l.startswith("M") and l[41] == "D")
        assert int(debit[42:55]) == 10000
        # montant négatif agrégé seul -> sens inversés sur le dossier 705
        charge_705 = next(l for l in par_paie["705"] if l[1:9] == "64133820")
        assert charge_705[41] == "C"
        assert inconnus == [("999999", "preca.xlsx")]   # centre inconnu + fichier
        assert attente == []

    def test_fichiers_disque_format_et_encodage(self, environnement):
        par_dossier, _ = generer_ecritures(environnement.sources, environnement)
        td, tc, deseq = ecrire_fichiers(par_dossier, environnement.dossier_sortie)
        assert td == tc and deseq == []
        brut = (open(f"{environnement.dossier_sortie}/704_ecriture_Quadra.txt", "rb").read())
        assert CRLF in brut and brut.endswith(CRLF)
        assert brut.count(b"\n") == brut.count(CRLF)            # CRLF strict
        lignes = brut.decode("cp1252").rstrip("\r\n").split("\r\n")
        assert all(len(l) in (39, 146) for l in lignes)
        assert all(l[0] in "MI" for l in lignes)

    def test_contre_passation_inverse_les_sens(self, environnement):
        normal, _ = generer_ecritures(environnement.sources, environnement)
        extourne, _ = generer_ecritures(environnement.sources, environnement, extourne=True)
        sens = lambda lots, cpt: next(
            l[41] for l in lots["704"] if l.startswith("M") and l[1:9] == cpt)
        assert sens(normal, "63388100") == "D" and sens(extourne, "63388100") == "C"
        assert sens(normal, "43785000") == "C" and sens(extourne, "43785000") == "D"
        # date d'extourne appliquée
        assert all(l[14:20] == "010626" for l in extourne["704"] if l.startswith("M"))

    def test_equilibre_global(self, environnement):
        par_dossier, _ = generer_ecritures(environnement.sources, environnement)
        par_paie, _, _ = generer_ecritures_paie(environnement.sources_paie, environnement)
        for d, lignes in par_paie.items():
            par_dossier[d].extend(lignes)
        for code, lignes in par_dossier.items():
            d, c = controler_equilibre(lignes)
            assert d == c, f"dossier {code} déséquilibré"


def test_agregation_cumule_un_seul_lot_par_dossier(tmp_path):
    """agreger=True : deux lignes du même dossier -> une seule écriture sur le cumul."""
    entree = tmp_path / "entree"
    sortie = tmp_path / "sortie"
    entree.mkdir()

    wb = Workbook()
    ws = wb.active
    ws.title = "CRE"
    for i, (code, montant) in enumerate(
            [("7704 - A", 1000.0), ("7704 - B", 500.0)], start=3):
        ws.cell(i, 1, code)
        ws.cell(i, 3, montant)
    wb.save(entree / "agg.xlsx")

    cfg = Configuration(
        dossier_entree=str(entree), dossier_sortie=str(sortie),
        analytique={"704": "770401"}, centre_vers_dossier={"770401": "704"},
        sources=[Source(
            fichier="agg.xlsx", feuille="CRE", ligne_debut=3,
            col_dossier="A", col_montant="C", extraire_code=True,
            compte_credit="40810000", compte_debit="62280000",
            libelle="AGG TEST", journal="OS", date_ecriture="310526",
            agreger=True)],
        sources_paie=[])

    par_dossier, sans_centre = generer_ecritures(cfg.sources, cfg)
    # une seule écriture = 2 lignes M (+1 ligne I), pas 4 lignes M
    assert sum(1 for l in par_dossier["704"] if l.startswith("M")) == 2
    debit = next(l for l in par_dossier["704"] if l.startswith("M") and l[41] == "D")
    assert int(debit[42:55]) == 150000          # 1000 + 500 = 1500 € cumulés
    assert sans_centre == []


def test_nettoyage_supprime_les_orphelins_ecriture_quadra(tmp_path):
    """Les *_ecriture_Quadra*.txt (y compris _contrepass) sont supprimés."""
    sortie = tmp_path / "sortie"
    sortie.mkdir()
    (sortie / "999_ecriture_Quadra.txt").write_text("orphelin", encoding="cp1252")
    (sortie / "704_ecriture_Quadra_contrepass.txt").write_text("x", encoding="cp1252")
    supprimes = nettoyer_sortie(str(sortie))
    assert not (sortie / "999_ecriture_Quadra.txt").exists()
    assert "999_ecriture_Quadra.txt" in supprimes
    assert "704_ecriture_Quadra_contrepass.txt" in supprimes   # le motif couvre le suffixe


def test_nettoyage_preserve_les_autres_fichiers(tmp_path):
    """notes.txt et donnees.xlsx ne correspondent pas au motif : conservés."""
    sortie = tmp_path / "sortie"
    sortie.mkdir()
    (sortie / "notes.txt").write_text("note", encoding="utf-8")
    (sortie / "donnees.xlsx").write_bytes(b"PK\x03\x04")
    (sortie / "736_ecriture_Quadra.txt").write_text("orphelin", encoding="cp1252")
    nettoyer_sortie(str(sortie))
    assert (sortie / "notes.txt").exists()
    assert (sortie / "donnees.xlsx").exists()
    assert not (sortie / "736_ecriture_Quadra.txt").exists()


def test_nettoyage_dossier_inexistant_sans_erreur(tmp_path):
    """Dossier de sortie pas encore créé : aucun fichier, aucune exception."""
    assert nettoyer_sortie(str(tmp_path / "pas_encore")) == []


def test_alias_dossier_route_vers_la_cible(tmp_path):
    """7736 -> dossier 723 (centre 772301) ; 7704 sans alias reste en 704."""
    entree = tmp_path / "entree"
    sortie = tmp_path / "sortie"
    entree.mkdir()

    wb = Workbook()
    ws = wb.active
    ws.title = "CRE"
    for i, (code, montant) in enumerate(
            [("7736 - ALIASEE", 100.0), ("7704 - NORMALE", 50.0)], start=3):
        ws.cell(i, 1, code)
        ws.cell(i, 3, montant)
    wb.save(entree / "alias.xlsx")

    cfg = Configuration(
        dossier_entree=str(entree), dossier_sortie=str(sortie),
        analytique={"723": "772301", "704": "770401"},
        centre_vers_dossier={}, sources_paie=[],
        alias_dossiers={"736": "723"},
        sources=[Source(
            fichier="alias.xlsx", feuille="CRE", ligne_debut=3,
            col_dossier="A", col_montant="C", extraire_code=True,
            compte_credit="40810000", compte_debit="62280000",
            libelle="ALIAS TEST", journal="OS", date_ecriture="310526")])

    par_dossier, sans_centre = generer_ecritures(cfg.sources, cfg)
    assert "723" in par_dossier and "736" not in par_dossier   # routé, pas de 736
    assert "704" in par_dossier                                # sans alias inchangé
    i_723 = next(l for l in par_dossier["723"] if l.startswith("I"))
    assert i_723[19:29].strip() == "772301"                    # centre de la cible
    assert sans_centre == []


def test_agregation_puis_ventilation(tmp_path):
    """agreger=True cumule, puis la ventilation répartit l'écriture agrégée."""
    entree = tmp_path / "entree"
    sortie = tmp_path / "sortie"
    entree.mkdir()

    wb = Workbook()
    ws = wb.active
    ws.title = "CRE"
    for i, (code, montant) in enumerate(
            [("704", 1000.0), ("704", 500.0)], start=2):
        ws.cell(i, 1, code)
        ws.cell(i, 3, montant)
    wb.save(entree / "aggvent.xlsx")

    cfg = Configuration(
        dossier_entree=str(entree), dossier_sortie=str(sortie),
        analytique={"704": "770401"}, centre_vers_dossier={}, sources_paie=[],
        sources=[Source(
            fichier="aggvent.xlsx", feuille="CRE", ligne_debut=2,
            col_dossier="A", col_montant="C",
            compte_credit="40810000", compte_debit="62280000",
            libelle="AGG VENT", journal="OS", date_ecriture="310526",
            agreger=True,
            ventilation={"704": [{"centre": "770401", "pourcent": 60.0},
                                 {"centre": "770402", "pourcent": 40.0}]})])

    par_dossier, _ = generer_ecritures(cfg.sources, cfg)
    # une seule écriture agrégée (2 M) ventilée en 2 lignes I sommant au cumul 1500 €
    assert sum(1 for l in par_dossier["704"] if l.startswith("M")) == 2
    i = [l for l in par_dossier["704"] if l.startswith("I")]
    assert len(i) == 2
    m_charge = next(int(l[42:55]) for l in par_dossier["704"]
                    if l.startswith("M") and l[1:9] == "62280000")
    assert m_charge == 150000
    assert sum(int(l[6:19]) for l in i) == m_charge            # 90000 + 60000


def _generer_avec_dates(tmp_path, lignes, **filtre):
    """lignes : liste de (code, montant, valeur_date) en colonnes A/C/E."""
    entree = tmp_path / "entree"
    sortie = tmp_path / "sortie"
    entree.mkdir()
    wb = Workbook()
    ws = wb.active
    ws.title = "CRE"
    for i, (code, montant, valeur_date) in enumerate(lignes, start=2):
        ws.cell(i, 1, code)
        ws.cell(i, 3, montant)
        if valeur_date is not None:
            ws.cell(i, 5, valeur_date)                 # colonne E
    wb.save(entree / "dates.xlsx")
    cfg = Configuration(
        dossier_entree=str(entree), dossier_sortie=str(sortie),
        analytique={"704": "770401"}, centre_vers_dossier={}, sources_paie=[],
        sources=[Source(
            fichier="dates.xlsx", feuille="CRE", ligne_debut=2,
            col_dossier="A", col_montant="C",
            compte_credit="40810000", compte_debit="62280000",
            libelle="DATE TEST", journal="OS", date_ecriture="310526",
            **filtre)])
    par_dossier, _ = generer_ecritures(cfg.sources, cfg)
    return par_dossier


def test_filtre_date_exclut_avant_la_borne_min(tmp_path):
    pd = _generer_avec_dates(tmp_path, [("704", 100.0, "20251231")],
                             col_date="E", date_min="20260101")
    assert "704" not in pd                             # 31/12/2025 hors période


def test_filtre_date_conserve_dans_la_periode(tmp_path):
    pd = _generer_avec_dates(tmp_path, [("704", 100.0, "20260531")],
                             col_date="E", date_min="20260101")
    assert "704" in pd


def test_filtre_date_max_exclut_apres_la_borne(tmp_path):
    pd = _generer_avec_dates(tmp_path, [("704", 100.0, "20260601")],
                             col_date="E", date_max="20260531")
    assert "704" not in pd                             # bornes incluses


def test_filtre_date_accepte_un_datetime(tmp_path):
    pd = _generer_avec_dates(tmp_path, [("704", 100.0, datetime(2025, 12, 31))],
                             col_date="E", date_min="20260101")
    assert "704" not in pd                             # datetime == chaîne équivalente


def test_sans_filtre_date_comportement_inchange(tmp_path):
    pd = _generer_avec_dates(tmp_path, [("704", 100.0, "20251231")])  # aucun filtre
    assert "704" in pd                                 # toutes les lignes conservées


def test_filtre_date_applique_avant_agregation(tmp_path):
    pd = _generer_avec_dates(
        tmp_path,
        [("704", 1000.0, "20260531"), ("704", 500.0, "20251231")],
        col_date="E", date_min="20260101", agreger=True)
    debit = next(l for l in pd["704"] if l.startswith("M") and l[41] == "D")
    assert int(debit[42:55]) == 100000                 # cumul = 1000 €, pas 1500


def _generer_pour_centres(tmp_path, lignes, *, analytique, centres_supp=None,
                          remap=None, ventilation=None):
    """Construit un classeur + une config et renvoie (par_dossier, centres_inconnus)."""
    entree = tmp_path / "entree"
    sortie = tmp_path / "sortie"
    entree.mkdir()
    wb = Workbook()
    ws = wb.active
    ws.title = "CRE"
    for i, (code, montant) in enumerate(lignes, start=2):
        ws.cell(i, 1, code)
        ws.cell(i, 3, montant)
    wb.save(entree / "centres.xlsx")
    centre_vers_dossier = {c: d for d, c in analytique.items()}
    centre_vers_dossier.update(centres_supp or {})
    cfg = Configuration(
        dossier_entree=str(entree), dossier_sortie=str(sortie),
        analytique=analytique, centre_vers_dossier=centre_vers_dossier,
        sources_paie=[], sources=[Source(
            fichier="centres.xlsx", feuille="CRE", ligne_debut=2,
            col_dossier="A", col_montant="C",
            compte_credit="40810000", compte_debit="62280000",
            libelle="CENTRE TEST", journal="OS", date_ecriture="310526",
            remap=remap or {}, ventilation=ventilation or {})])
    centres_inconnus = []
    par_dossier, _ = generer_ecritures(cfg.sources, cfg, centres_inconnus=centres_inconnus)
    return par_dossier, centres_inconnus


def test_centre_dans_analytique_accepte_sans_signalement(tmp_path):
    par, sig = _generer_pour_centres(tmp_path, [("790", 100.0)],
                                     analytique={"790": "179101"})
    assert "790" in par and sig == []


def test_centre_dans_centres_supplementaires_accepte(tmp_path):
    # le code lu 179102 (remap -> dossier 790) sert de centre ; il est connu
    par, sig = _generer_pour_centres(
        tmp_path, [("179102", 100.0)],
        analytique={"790": "179101"}, centres_supp={"179102": "790"},
        remap={"179102": "790"})
    assert sig == []


def test_centre_dans_ventilation_accepte(tmp_path):
    par, sig = _generer_pour_centres(
        tmp_path, [("790", 100.0)], analytique={"790": "179101"},
        ventilation={"790": [{"centre": "179103", "pourcent": 100.0}]})
    assert sig == []


def test_centre_inconnu_signale_sans_bloquer(tmp_path):
    # 179104 n'est ni dans analytique, ni supplémentaire, ni ventilé -> signalé
    par, sig = _generer_pour_centres(
        tmp_path, [("179104", 100.0)],
        analytique={"790": "179101"}, remap={"179104": "790"})
    assert "790" in par                                # écriture bien produite
    assert any(l.startswith("I") for l in par["790"])  # ligne I produite quand même
    assert len(sig) == 1
    centre, dossier, libelle, fichier = sig[0]
    assert centre == "179104" and dossier == "790"
    assert libelle == "CENTRE TEST" and fichier == "centres.xlsx"


def _generer_paie_centres(tmp_path, fichiers):
    """fichiers : liste de (nom_fichier, [(centre, montant), ...]).

    Renvoie la liste centres_inconnus de generer_ecritures_paie.
    """
    entree = tmp_path / "entree"
    sortie = tmp_path / "sortie"
    entree.mkdir()
    sources = []
    for nom, lignes in fichiers:
        wb = Workbook()
        ws = wb.active
        ws.title = "Feuil1"
        for i, (centre, montant) in enumerate(lignes, start=2):
            ws.cell(i, 4, centre)                      # colonne D : centre de coût
            ws.cell(i, 14, montant)                    # colonne N : composante
        wb.save(entree / nom)
        sources.append(SourcePaie(
            fichier=nom, feuille="Feuil1", ligne_debut=2, col_centre="D",
            journal="OS", date_ecriture="310526", contre_passation=None,
            composantes=[Composante(col="N", compte_debit="64133820",
                                    compte_credit="42822000", libelle="PRIME")]))
    cfg = Configuration(
        dossier_entree=str(entree), dossier_sortie=str(sortie),
        analytique={"704": "770401"}, centre_vers_dossier={"770401": "704"},
        sources=[], sources_paie=sources)
    _, centres_inconnus, _ = generer_ecritures_paie(cfg.sources_paie, cfg)
    return centres_inconnus


def test_centre_paie_non_route_signale_avec_son_fichier(tmp_path):
    inconnus = _generer_paie_centres(tmp_path, [
        ("EDU PRIME 310526.xlsx", [("770401", 100.0), ("999999", 50.0)])])
    assert inconnus == [("999999", "EDU PRIME 310526.xlsx")]


def test_deux_fichiers_paie_centres_non_routes_distincts(tmp_path):
    inconnus = _generer_paie_centres(tmp_path, [
        ("A.xlsx", [("999998", 10.0)]),
        ("B.xlsx", [("999999", 20.0)])])
    assert inconnus == [("999998", "A.xlsx"), ("999999", "B.xlsx")]


def test_tous_centres_paie_routes_aucun_signalement(tmp_path):
    inconnus = _generer_paie_centres(tmp_path, [("A.xlsx", [("770401", 100.0)])])
    assert inconnus == []


def _config_numero_piece(tmp_path, numero_piece_global, numero_piece_source):
    entree = tmp_path / "entree"
    sortie = tmp_path / "sortie"
    entree.mkdir()
    wb = Workbook()
    ws = wb.active
    ws.title = "CRE"
    ws.cell(2, 1, "704")
    ws.cell(2, 3, 100.0)
    wb.save(entree / "p.xlsx")
    source = Source(
        fichier="p.xlsx", feuille="CRE", ligne_debut=2, col_dossier="A",
        col_montant="C", compte_credit="40810000", compte_debit="62280000",
        libelle="X", journal="OS", date_ecriture="310526",
        numero_piece=numero_piece_source)
    cfg = Configuration(
        dossier_entree=str(entree), dossier_sortie=str(sortie),
        analytique={"704": "770401"}, centre_vers_dossier={}, sources_paie=[],
        sources=[source], numero_piece=numero_piece_global)
    par, _ = generer_ecritures(cfg.sources, cfg)
    return next(l for l in par["704"] if l.startswith("M"))


def test_numero_piece_source_prime_sur_la_globale(tmp_path):
    m = _config_numero_piece(tmp_path, numero_piece_global="GLOBAL",
                             numero_piece_source="SOURCE")
    assert m[99:107] == "SOURCE  "


def test_numero_piece_globale_quand_source_absente(tmp_path):
    m = _config_numero_piece(tmp_path, numero_piece_global="GLOBAL",
                             numero_piece_source=None)
    assert m[99:107] == "GLOBAL  "


def _run_cli_numero_piece(tmp_path, incremental):
    """Lance cli.main sur une config minimale et renvoie les lignes M produites
    (passe normale + contre-passation)."""
    entree = tmp_path / "entree"
    sortie = tmp_path / "sortie"
    if not entree.exists():
        entree.mkdir()
    wb = Workbook()
    ws = wb.active
    ws.title = "CRE"
    ws.cell(2, 1, "704")
    ws.cell(2, 3, 100.0)
    wb.save(entree / "p.xlsx")
    cfg = (
        f'dossier_entree: "{entree.as_posix()}"\n'
        f'dossier_sortie: "{sortie.as_posix()}"\n'
        'numero_piece: "IMPORT"\n'
        + ('numero_piece_incremental: true\n' if incremental else '')
        + 'analytique:\n  "704": "770401"\n'
        'sources:\n'
        '  - fichier: "p.xlsx"\n'
        '    feuille: "CRE"\n'
        '    ligne_debut: 2\n'
        '    col_dossier: "A"\n'
        '    col_montant: "C"\n'
        '    compte_credit: "40810000"\n'
        '    compte_debit: "62280000"\n'
        '    libelle: "X"\n'
        '    journal: "OS"\n'
        '    date_ecriture: "310526"\n'
        '    contre_passation: "010626"\n'
        'sources_paie: []\n'
    )
    chemin = tmp_path / "cfg.yaml"
    chemin.write_text(cfg, encoding="utf-8")
    cli_main(["--config", str(chemin)])
    mlines = []
    for nom in ("704_ecriture_Quadra.txt", "704_ecriture_Quadra_contrepass.txt"):
        p = sortie / nom
        if p.exists():
            mlines += [l.decode("cp1252")
                       for l in p.read_bytes().split(b"\r\n") if l[:1] == b"M"]
    return mlines


def test_numero_piece_incremental_deux_runs_successifs(tmp_path):
    m1 = _run_cli_numero_piece(tmp_path, incremental=True)
    assert m1 and all(l[99:107] == "IMPORT01" for l in m1)   # tout le run #1
    assert all(len(l) == 146 for l in m1)
    m2 = _run_cli_numero_piece(tmp_path, incremental=True)
    assert m2 and all(l[99:107] == "IMPORT02" for l in m2)   # run #2 incrémenté


def test_numero_piece_fixe_sans_option_incremental(tmp_path):
    m = _run_cli_numero_piece(tmp_path, incremental=False)
    assert m and all(l[99:107] == "IMPORT  " for l in m)     # fixe, comportement inchangé


def _generer_paie_doublons(tmp_path, fichiers):
    """fichiers : liste de (nom, [(matricule, centre, montant), ...]).

    Matricule en colonne G, centre en D, montant en N. Renvoie la liste doublons.
    """
    entree = tmp_path / "entree"
    sortie = tmp_path / "sortie"
    entree.mkdir()
    sources = []
    for nom, lignes in fichiers:
        wb = Workbook()
        ws = wb.active
        ws.title = "Feuil1"
        for i, (matricule, centre, montant) in enumerate(lignes, start=2):
            if matricule is not None:
                ws.cell(i, 7, matricule)               # colonne G : matricule
            ws.cell(i, 4, centre)                      # colonne D : centre
            ws.cell(i, 14, montant)                    # colonne N : composante
        wb.save(entree / nom)
        sources.append(SourcePaie(
            fichier=nom, feuille="Feuil1", ligne_debut=2, col_centre="D",
            journal="OS", date_ecriture="310526", contre_passation=None,
            composantes=[Composante(col="N", compte_debit="64133820",
                                    compte_credit="42822000", libelle="PRIME")]))
    cfg = Configuration(
        dossier_entree=str(entree), dossier_sortie=str(sortie),
        analytique={"704": "770401"}, centre_vers_dossier={"770401": "704"},
        sources=[], sources_paie=sources)
    doublons = []
    generer_ecritures_paie(cfg.sources_paie, cfg, doublons=doublons)
    return doublons


def test_doublon_meme_matricule_centre_dans_deux_fichiers(tmp_path):
    d = _generer_paie_doublons(tmp_path, [
        ("paieA.xlsx", [("M001", "770401", 100.0)]),
        ("paieB.xlsx", [("M001", "770401", 100.0)])])
    assert d == [("M001", "770401", ("paieA.xlsx", "paieB.xlsx"))]


def test_meme_matricule_centres_differents_pas_de_doublon(tmp_path):
    d = _generer_paie_doublons(tmp_path, [
        ("paieA.xlsx", [("M001", "770401", 100.0)]),
        ("paieB.xlsx", [("M001", "770501", 100.0)])])
    assert d == []                                     # salarié réparti : normal


def test_aucun_doublon_aucune_detection(tmp_path):
    d = _generer_paie_doublons(tmp_path, [
        ("paieA.xlsx", [("M001", "770401", 100.0)]),
        ("paieB.xlsx", [("M002", "770401", 100.0)])])
    assert d == []


def test_matricule_absent_ignore(tmp_path):
    d = _generer_paie_doublons(tmp_path, [
        ("paieA.xlsx", [(None, "770401", 100.0)]),
        ("paieB.xlsx", [(None, "770401", 100.0)])])
    assert d == []


def _generer_paie_cols(tmp_path, fichiers, col_matricule="G"):
    """fichiers : liste de (nom, [(g, h, centre, montant), ...]) — G, H, D, N."""
    entree = tmp_path / "entree"
    sortie = tmp_path / "sortie"
    if not entree.exists():
        entree.mkdir()
    sources = []
    for nom, lignes in fichiers:
        wb = Workbook()
        ws = wb.active
        ws.title = "Feuil1"
        for i, (g, h, centre, montant) in enumerate(lignes, start=2):
            if g is not None:
                ws.cell(i, 7, g)                       # G : catégorie
            if h is not None:
                ws.cell(i, 8, h)                       # H : matricule (STC)
            ws.cell(i, 4, centre)                      # D : centre
            ws.cell(i, 14, montant)                    # N : composante
        wb.save(entree / nom)
        sources.append(SourcePaie(
            fichier=nom, feuille="Feuil1", ligne_debut=2, col_centre="D",
            journal="OS", date_ecriture="310526", contre_passation=None,
            col_matricule=col_matricule,
            composantes=[Composante(col="N", compte_debit="64133820",
                                    compte_credit="42822000", libelle="PRIME")]))
    cfg = Configuration(
        dossier_entree=str(entree), dossier_sortie=str(sortie),
        analytique={"704": "770401"}, centre_vers_dossier={"770401": "704"},
        sources=[], sources_paie=sources)
    doublons = []
    generer_ecritures_paie(cfg.sources_paie, cfg, doublons=doublons)
    return doublons


def test_doublon_matricule_en_colonne_h(tmp_path):
    # G = catégorie identique (à ne pas lire), H = vrai matricule
    d = _generer_paie_cols(tmp_path, [
        ("A.xlsx", [("Medical", "M001", "770401", 100.0)]),
        ("B.xlsx", [("Medical", "M001", "770401", 100.0)])], col_matricule="H")
    assert d == [("M001", "770401", ("A.xlsx", "B.xlsx"))]


def test_detection_desactivee_quand_col_matricule_none(tmp_path):
    d = _generer_paie_cols(tmp_path, [
        ("A.xlsx", [("Medical", "M001", "770401", 100.0)]),
        ("B.xlsx", [("Medical", "M001", "770401", 100.0)])], col_matricule=None)
    assert d == []                                     # détection désactivée


def _generer_avec_entete(tmp_path, entete_cellules, entete_attendu, ligne_entete=None):
    """Écrit un en-tête en ligne 1 et une donnée en ligne 2 (ligne_debut=2)."""
    entree = tmp_path / "entree"
    sortie = tmp_path / "sortie"
    entree.mkdir()
    wb = Workbook()
    ws = wb.active
    ws.title = "CRE"
    for col, val in entete_cellules.items():
        ws[f"{col}1"] = val
    ws["A2"] = "704"
    ws["C2"] = 100.0
    wb.save(entree / "p.xlsx")
    src = Source(
        fichier="p.xlsx", feuille="CRE", ligne_debut=2, col_dossier="A",
        col_montant="C", compte_credit="40810000", compte_debit="62280000",
        libelle="X", journal="OS", date_ecriture="310526",
        entete_attendu=entete_attendu, ligne_entete=ligne_entete)
    cfg = Configuration(
        dossier_entree=str(entree), dossier_sortie=str(sortie),
        analytique={"704": "770401"}, centre_vers_dossier={}, sources_paie=[],
        sources=[src])
    return generer_ecritures(cfg.sources, cfg)


def test_entete_conforme_traitement_normal(tmp_path):
    par, _ = _generer_avec_entete(
        tmp_path, {"A": "Dossier", "C": "Montant"},
        entete_attendu={"A": "Dossier", "C": "Montant"})
    assert "704" in par


def test_entete_non_conforme_erreur_bloquante(tmp_path):
    with pytest.raises(EnteteInvalide) as exc:
        _generer_avec_entete(
            tmp_path, {"A": "AUTRE CHOSE", "C": "Montant"},
            entete_attendu={"A": "Dossier", "C": "Montant"})
    msg = str(exc.value)
    assert "p.xlsx" in msg and "A" in msg          # fichier + colonne
    assert "Dossier" in msg and "AUTRE CHOSE" in msg   # attendu + trouvé


def test_entete_comparaison_insensible_casse_et_espaces(tmp_path):
    par, _ = _generer_avec_entete(
        tmp_path, {"A": "  DOSSIER ", "C": "montant"},
        entete_attendu={"A": "Dossier", "C": "Montant"})
    assert "704" in par                            # toléré


def test_sans_entete_attendu_comportement_inchange(tmp_path):
    par, _ = _generer_avec_entete(
        tmp_path, {"A": "peu importe"}, entete_attendu={})
    assert "704" in par


def _ecrire_cfg_reference(tmp_path, entree, sortie, ref, montant, avec_reference):
    wb = Workbook()
    ws = wb.active
    ws.title = "CRE"
    ws["A2"] = "704"
    ws["C2"] = montant
    wb.save(entree / "p.xlsx")
    lignes = [f'dossier_entree: "{entree.as_posix()}"',
              f'dossier_sortie: "{sortie.as_posix()}"']
    if avec_reference:
        lignes.append(f'dossier_reference: "{ref.as_posix()}"')
    lignes += ['analytique:', '  "704": "770401"', 'sources:',
               '  - fichier: "p.xlsx"', '    feuille: "CRE"', '    ligne_debut: 2',
               '    col_dossier: "A"', '    col_montant: "C"',
               '    compte_credit: "40810000"', '    compte_debit: "62280000"',
               '    libelle: "Charge"', '    journal: "OS"',
               '    date_ecriture: "310526"', 'sources_paie: []']
    chemin = tmp_path / "cfg.yaml"
    chemin.write_text("\n".join(lignes) + "\n", encoding="utf-8")
    return chemin


def test_dossier_reference_du_yaml_sans_option_cli(tmp_path):
    """Cas Lancer.bat : la clé config dossier_reference suffit (sans --reference)."""
    entree = tmp_path / "entree"
    sortie = tmp_path / "sortie"
    ref = tmp_path / "reference"
    entree.mkdir()
    # 1er run sans référence -> on en fait la version de référence
    cli_main(["--config",
              str(_ecrire_cfg_reference(tmp_path, entree, sortie, ref, 100.0, False))])
    shutil.copytree(sortie, ref)
    # 2e run : dossier_reference dans le YAML, AUCUN argument --reference
    cli_main(["--config",
              str(_ecrire_cfg_reference(tmp_path, entree, sortie, ref, 150.0, True))])
    csvs = list(sortie.glob("diff_situation_*.csv"))
    assert len(csvs) == 1                              # diff généré via la clé config seule
    lignes = csvs[0].read_text(encoding="utf-8-sig").splitlines()
    assert lignes[0].startswith("Type;Dossier")
    assert any("MONTANT_MODIFIE;704" in l for l in lignes)


def test_archiver_entree_zip_contient_les_fichiers(tmp_path):
    import zipfile
    entree = tmp_path / "entree"
    arch = tmp_path / "archives"
    entree.mkdir()
    (entree / "a.xlsx").write_bytes(b"AAA")
    (entree / "b.xlsx").write_bytes(b"BBB")
    chemin = archiver_entree(str(entree), str(arch), "20260617_120000")
    assert os.path.basename(chemin) == "entree_20260617_120000.zip"   # motif horodaté
    with zipfile.ZipFile(chemin) as z:
        assert sorted(z.namelist()) == ["a.xlsx", "b.xlsx"]


def test_archiver_entree_deux_archives_distinctes(tmp_path):
    entree = tmp_path / "entree"
    arch = tmp_path / "archives"
    entree.mkdir()
    (entree / "a.xlsx").write_bytes(b"AAA")
    c1 = archiver_entree(str(entree), str(arch), "20260617_120000")
    c2 = archiver_entree(str(entree), str(arch), "20260617_120001")
    assert c1 != c2 and os.path.exists(c1) and os.path.exists(c2)


def test_archiver_entree_vide_pas_de_zip(tmp_path):
    entree = tmp_path / "entree"
    entree.mkdir()
    assert archiver_entree(str(entree), str(tmp_path / "archives"), "20260617_120000") is None


def test_archiver_entree_absent_pas_de_zip(tmp_path):
    assert archiver_entree(str(tmp_path / "absent"), str(tmp_path / "archives"),
                           "20260617_120000") is None


def _cfg_minimal_archive(tmp_path, entree, sortie, lignes_extra=()):
    wb = Workbook()
    ws = wb.active
    ws.title = "CRE"
    ws["A2"] = "704"
    ws["C2"] = 100.0
    wb.save(entree / "p.xlsx")
    lignes = [f'dossier_entree: "{entree.as_posix()}"',
              f'dossier_sortie: "{sortie.as_posix()}"', *lignes_extra,
              'analytique:', '  "704": "770401"', 'sources:',
              '  - fichier: "p.xlsx"', '    feuille: "CRE"', '    ligne_debut: 2',
              '    col_dossier: "A"', '    col_montant: "C"',
              '    compte_credit: "40810000"', '    compte_debit: "62280000"',
              '    libelle: "Charge"', '    journal: "OS"', '    date_ecriture: "310526"',
              'sources_paie: []']
    chemin = tmp_path / "cfg.yaml"
    chemin.write_text("\n".join(lignes) + "\n", encoding="utf-8")
    return chemin


def test_cli_archive_entree_si_dossier_archives(tmp_path):
    import zipfile
    entree = tmp_path / "entree"
    sortie = tmp_path / "sortie"
    arch = tmp_path / "archives"
    entree.mkdir()
    chemin = _cfg_minimal_archive(tmp_path, entree, sortie,
                                  [f'dossier_archives: "{arch.as_posix()}"'])
    cli_main(["--config", str(chemin)])
    zips = list(arch.glob("entree_*.zip"))
    assert len(zips) == 1
    with zipfile.ZipFile(zips[0]) as z:
        assert "p.xlsx" in z.namelist()


def test_cli_aucun_archivage_sans_dossier_archives(tmp_path):
    entree = tmp_path / "entree"
    sortie = tmp_path / "sortie"
    arch = tmp_path / "archives"
    entree.mkdir()
    cli_main(["--config", str(_cfg_minimal_archive(tmp_path, entree, sortie))])
    assert not arch.exists()                           # pas d'archivage par défaut


def test_source_correspond_sous_chaine_et_glob():
    assert source_correspond("CRE PRECA 310526.xlsx", "PRECA")       # sous-chaîne
    assert source_correspond("CRE PRECA 310526.xlsx", "*preca*")     # glob + casse
    assert source_correspond("produits.xlsx", "*.xlsx")              # glob extension
    assert not source_correspond("charges.xlsx", "produits")


def _cfg_deux_sources(tmp_path):
    entree = tmp_path / "entree"
    sortie = tmp_path / "sortie"
    entree.mkdir()
    for nom, code in [("produits.xlsx", "704"), ("charges.xlsx", "705")]:
        wb = Workbook()
        ws = wb.active
        ws.title = "CRE"
        ws["A2"] = code
        ws["C2"] = 100.0
        wb.save(entree / nom)
    lignes = [f'dossier_entree: "{entree.as_posix()}"',
              f'dossier_sortie: "{sortie.as_posix()}"',
              'analytique:', '  "704": "770401"', '  "705": "770501"', 'sources:']
    for nom in ("produits.xlsx", "charges.xlsx"):
        lignes += [f'  - fichier: "{nom}"', '    feuille: "CRE"', '    ligne_debut: 2',
                   '    col_dossier: "A"', '    col_montant: "C"',
                   '    compte_credit: "40810000"', '    compte_debit: "62280000"',
                   f'    libelle: "Lib {nom[:4]}"', '    journal: "OS"',
                   '    date_ecriture: "310526"']
    lignes.append('sources_paie: []')
    chemin = tmp_path / "cfg.yaml"
    chemin.write_text("\n".join(lignes) + "\n", encoding="utf-8")
    return chemin, sortie


def _fichiers_sortie(sortie):
    return {p.name for p in sortie.glob("*_ecriture_Quadra.txt")}


def test_source_unique_restreint_le_perimetre(tmp_path):
    chemin, sortie = _cfg_deux_sources(tmp_path)
    cli_main(["--config", str(chemin), "--source", "produits"])
    assert _fichiers_sortie(sortie) == {"704_ecriture_Quadra.txt"}   # source ciblée seule


def test_sans_source_toutes_les_sources(tmp_path):
    chemin, sortie = _cfg_deux_sources(tmp_path)
    cli_main(["--config", str(chemin)])
    assert _fichiers_sortie(sortie) == {"704_ecriture_Quadra.txt",
                                        "705_ecriture_Quadra.txt"}


def test_perimetre_restreint_reste_equilibre(tmp_path):
    chemin, sortie = _cfg_deux_sources(tmp_path)
    assert cli_main(["--config", str(chemin), "--source", "produits"]) == 0
    data = (sortie / "704_ecriture_Quadra.txt").read_bytes().decode("cp1252")
    m = [l for l in data.split("\r\n") if l.startswith("M")]
    debit = sum(int(l[42:55]) for l in m if l[41] == "D")
    credit = sum(int(l[42:55]) for l in m if l[41] == "C")
    assert debit == credit and debit > 0


def test_output_dossier_distinct_preserve_la_generation_complete(tmp_path):
    chemin, sortie = _cfg_deux_sources(tmp_path)
    cli_main(["--config", str(chemin)])                # génération complète
    out = tmp_path / "sortie_produits"
    cli_main(["--config", str(chemin), "--source", "produits", "--output", str(out)])
    assert (sortie / "705_ecriture_Quadra.txt").exists()        # complète préservée
    assert _fichiers_sortie(out) == {"704_ecriture_Quadra.txt"}  # restreint à part


PRIME_FICHIER = "CRE SIEGE PRIME DECENTRALISEE 310526.xlsx"


def _cfg_prime_decentralisee(tmp_path, centre="770401", brut=1000.0,
                             header_m="Centre de coût new"):
    """Source de paie « prime décentralisée » : charges calculées à partir du brut
    (col S) via un taux par composante ; charges absentes du fichier."""
    entree = tmp_path / "entree"
    sortie = tmp_path / "sortie"
    if not entree.exists():
        entree.mkdir()
    wb = Workbook()
    ws = wb.active
    ws.title = "Feuil1"
    ws["M2"] = header_m                                # en-tête ligne 2
    ws["G2"] = "Matricule"
    ws["S2"] = "PRIME DECENTRALISEE"
    ws["M3"] = centre                                  # données ligne 3
    ws["G3"] = "M001"
    ws["S3"] = brut
    wb.save(entree / PRIME_FICHIER)
    return Configuration(
        dossier_entree=str(entree), dossier_sortie=str(sortie),
        analytique={"704": "770401"}, centre_vers_dossier={"770401": "704"},
        sources=[], sources_paie=[SourcePaie(
            fichier=PRIME_FICHIER, feuille="Feuil1", ligne_debut=3,
            col_centre="M", col_matricule="G", journal="OS", date_ecriture="310526",
            entete_attendu={"M": "Centre de coût new", "G": "Matricule",
                            "S": "PRIME DECENTRALISEE"},
            composantes=[
                Composante(col="S", compte_debit="64133840",
                           compte_credit="42824000", libelle="Prime decentr 0526"),
                Composante(col="S", taux=0.3592, compte_debit="64585000",
                           compte_credit="43824000", libelle="Ch soc decentr 0526"),
                Composante(col="S", taux=0.0987, compte_debit="63185000",
                           compte_credit="44824000", libelle="Ch fisc decentr 0526"),
            ])])


def _montant_m(lignes, compte, sens):
    return next(int(l[42:55]) for l in lignes
                if l.startswith("M") and l[1:9] == compte and l[41] == sens)


def test_prime_decentralisee_charges_calculees(tmp_path):
    cfg = _cfg_prime_decentralisee(tmp_path, brut=1000.0)
    par, inconnus, attente = generer_ecritures_paie(cfg.sources_paie, cfg)
    lignes = par["704"]
    assert _montant_m(lignes, "64133840", "D") == 100000   # brut 1000,00 €
    assert _montant_m(lignes, "64585000", "D") == 35920    # 1000 × 0,3592
    assert _montant_m(lignes, "63185000", "D") == 9870     # 1000 × 0,0987
    assert inconnus == [] and attente == []


def test_prime_decentralisee_total_equilibre(tmp_path):
    cfg = _cfg_prime_decentralisee(tmp_path, brut=1234.56)
    par, _, _ = generer_ecritures_paie(cfg.sources_paie, cfg)
    d, c = controler_equilibre(par["704"])
    assert d == c and d > 0                                 # chaque composante équilibrée


def test_prime_decentralisee_entete_non_conforme_bloque(tmp_path):
    cfg = _cfg_prime_decentralisee(tmp_path, header_m="MAUVAIS LIBELLE")
    with pytest.raises(EnteteInvalide):
        generer_ecritures_paie(cfg.sources_paie, cfg)


def test_prime_decentralisee_centre_inconnu_signale(tmp_path):
    cfg = _cfg_prime_decentralisee(tmp_path, centre="999999")
    _, inconnus, _ = generer_ecritures_paie(cfg.sources_paie, cfg)
    assert inconnus == [("999999", PRIME_FICHIER)]
