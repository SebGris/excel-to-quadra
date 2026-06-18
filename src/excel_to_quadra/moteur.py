# -*- coding: utf-8 -*-
"""Moteur de génération des écritures Quadra.

Chaque écriture est produite en paire équilibrée (un débit, un crédit du même
montant). Une ligne analytique (type I) est ajoutée immédiatement après la part
de charge ou de produit (comptes de classe 6 ou 7), ventilée à 100 % vers le
centre du dossier.

Règles de gestion :
  - montant négatif (régularisation)  -> sens débit/crédit inversés ;
  - extourne (contre-passation)        -> sens inversés à la date d'extourne ;
  - extourne d'un montant négatif      -> double inversion = sens d'origine ;
  - centre analytique inconnu          -> ligne M produite, absence signalée.
"""

import fnmatch
import glob
import os
import zipfile
from collections import defaultdict
from decimal import ROUND_HALF_UP, Decimal
from typing import Dict, List, Optional, Tuple

#: Motif des fichiers générés (un par dossier, + variante _contrepass).
MOTIF_SORTIE = "*_ecriture_Quadra*.txt"

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from .config import Composante, Configuration, Source, SourcePaie
from .format_quadra import formater_ligne_i, formater_ligne_m
from .normalisation import lire_montant, normaliser_code

ParDossier = Dict[str, List[str]]


class EnteteInvalide(Exception):
    """Levée quand l'en-tête réel d'un classeur ne correspond pas à `entete_attendu`."""


def _est_charge_ou_produit(compte: str) -> bool:
    return str(compte)[:1] in ("6", "7")


def source_correspond(fichier: str, motif: str) -> bool:
    """True si le nom de fichier correspond au motif — sous-chaîne **ou** motif
    glob (`*`, `?`) —, insensible à la casse. Sert à restreindre le périmètre de
    génération à une (ou des) source(s) précise(s)."""
    f, m = str(fichier).lower(), str(motif).lower()
    return m in f or fnmatch.fnmatchcase(f, m)


def _verifier_entete(ws, entete_attendu: Dict[str, str], ligne_entete: int,
                     fichier: str) -> None:
    """Vérifie chaque libellé d'en-tête déclaré ; lève EnteteInvalide en cas d'écart.

    Comparaison souple : insensible à la casse et aux espaces de début/fin.
    """
    for colonne, attendu in entete_attendu.items():
        trouve = ws.cell(ligne_entete, column_index_from_string(colonne)).value
        trouve_txt = "" if trouve is None else str(trouve)
        if trouve_txt.strip().lower() != str(attendu).strip().lower():
            raise EnteteInvalide(
                f"En-tête inattendu dans « {fichier} », colonne {colonne} "
                f"(ligne {ligne_entete}) : attendu « {attendu} », trouvé « {trouve_txt} ». "
                f"Le fichier ne correspond pas à ce que la configuration croit lire.")


def _normaliser_date(valeur) -> Optional[str]:
    """Renvoie la date en chaîne AAAAMMJJ (8 chiffres), ou None si inexploitable.

    Deux formes possibles selon l'export : une chaîne « 20260531 » déjà au bon
    format, ou un datetime/date openpyxl (alors formaté en %Y%m%d).
    """
    if valeur is None:
        return None
    if hasattr(valeur, "strftime"):            # datetime / date
        return valeur.strftime("%Y%m%d")
    texte = str(valeur).strip()
    return texte if len(texte) == 8 and texte.isdigit() else None


def ajouter_ecriture_pair(par_dossier: ParDossier, dossier: str, journal: str,
                          date: str, libelle: str, compte_credit: str,
                          compte_debit: str, montant: float,
                          centre: Optional[str],
                          sans_centre: Optional[list],
                          extourne: bool = False, facteur: float = 1.0,
                          ventilation: Optional[List[dict]] = None,
                          centres_connus: Optional[set] = None,
                          centres_inconnus: Optional[list] = None,
                          fichier: Optional[str] = None,
                          numero_piece: Optional[str] = None) -> None:
    """Ajoute une écriture équilibrée (M crédit + M débit) et sa/ses ligne(s) I.

    Sans ventilation : une ligne I à 100 % vers `centre`. Avec ventilation (liste
    de {centre, pourcent}) : une ligne I par centre, la dernière recevant le solde
    pour que la somme des lignes I égale exactement la ligne M.

    Si `centres_connus` est fourni, chaque centre produit absent de cet ensemble
    est mémorisé dans `centres_inconnus` (avertissement, sans bloquer l'écriture).
    """
    def _verifier(centre_produit):
        if (centres_connus is not None and centres_inconnus is not None
                and centre_produit not in centres_connus):
            centres_inconnus.append((centre_produit, dossier, libelle, fichier))

    montant = round(montant * facteur, 2)      # prorata avant calcul du sens
    inverser = extourne ^ (montant < 0)        # extourne et négatif se cumulent
    montant = abs(montant)
    if inverser:                               # l'ex-débit passe au crédit
        legs = [(compte_debit, "C"), (compte_credit, "D")]
    else:
        legs = [(compte_credit, "C"), (compte_debit, "D")]
    # part de charge/produit (classe 6/7) écrite en premier, suivie de sa ligne I
    legs.sort(key=lambda x: 0 if _est_charge_ou_produit(x[0]) else 1)
    for compte, sens in legs:
        par_dossier[dossier].append(
            formater_ligne_m(compte, journal, date, libelle, sens, montant, numero_piece))
        if not _est_charge_ou_produit(compte):
            continue
        if ventilation:                        # ventilation prioritaire sur le centre
            # Arrondi commercial (Decimal/ROUND_HALF_UP) comme pour les centimes :
            # jamais d'arithmétique flottante sur les montants comptables.
            reste = Decimal(str(montant))
            dernier = len(ventilation) - 1
            for i, v in enumerate(ventilation):
                if i < dernier:
                    part = (Decimal(str(montant)) * Decimal(str(v["pourcent"])) / 100
                            ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                    reste -= part
                else:
                    part = reste               # le solde garantit somme(I) == M
                par_dossier[dossier].append(formater_ligne_i(v["centre"], part, v["pourcent"]))
                _verifier(v["centre"])
        elif centre:
            par_dossier[dossier].append(formater_ligne_i(centre, montant))
            _verifier(centre)
        elif sans_centre is not None:
            sans_centre.append((dossier, libelle, fichier))   # fichier pour la traçabilité


def generer_ecritures(sources: List[Source], cfg: Configuration,
                      extourne: bool = False,
                      centres_inconnus: Optional[list] = None) -> Tuple[ParDossier, list]:
    """Traite les sources « une ligne = un établissement ».

    Renvoie (par_dossier, sans_centre) où sans_centre liste les triplets
    (dossier, libellé, fichier) dont la ligne analytique n'a pas pu être générée.

    Si `centres_inconnus` (liste) est fourni, les centres produits absents de
    `cfg.centres_connus()` y sont mémorisés (avertissement, sans bloquer).
    """
    par_dossier: ParDossier = defaultdict(list)
    sans_centre: list = []
    connus = cfg.centres_connus() if centres_inconnus is not None else None
    cache_wb = {}

    for src in sources:
        if extourne and not src.contre_passation:
            continue
        chemin = os.path.join(cfg.dossier_entree, src.fichier)
        if chemin not in cache_wb:
            cache_wb[chemin] = load_workbook(chemin, data_only=True)
        ws = cache_wb[chemin][src.feuille]

        if src.entete_attendu:                       # contrôle de structure (bloquant)
            _verifier_entete(ws, src.entete_attendu,
                             src.ligne_entete or (src.ligne_debut - 1), src.fichier)

        c_dossier = column_index_from_string(src.col_dossier)
        c_montant = column_index_from_string(src.col_montant)
        date = src.contre_passation if extourne else src.date_ecriture

        # Filtre de dates : actif uniquement si une colonne date ET au moins une
        # borne sont renseignées. Comparaison lexicographique d'AAAAMMJJ (= ordre
        # chronologique). Appliqué avant l'agrégation pour ne cumuler que la période.
        c_date = column_index_from_string(src.col_date) if src.col_date else None
        filtre_actif = c_date is not None and (src.date_min or src.date_max)
        numero_piece = src.numero_piece or cfg.numero_piece   # surcharge source > globale

        # En mode agrégé, on cumule (dossier, centre) -> montant et on émet
        # une seule écriture par dossier après la boucle, au lieu de ligne à ligne.
        cumul: Dict[Tuple[str, Optional[str]], float] = defaultdict(float)

        for r in range(src.ligne_debut, ws.max_row + 1):
            code_brut = normaliser_code(ws.cell(r, c_dossier).value,
                                        src.extraire_code, src.strip_zeros)
            montant = lire_montant(ws.cell(r, c_montant).value)
            if code_brut is None or montant is None:
                continue
            if filtre_actif:
                d = _normaliser_date(ws.cell(r, c_date).value)
                if (d is None or (src.date_min and d < src.date_min)
                        or (src.date_max and d > src.date_max)):
                    continue                   # date absente/hors bornes : ignorée
            # Alias : le dossier lu est remplacé par sa cible (centre résolu sur
            # la cible) — contrairement au remap, le code lu n'est pas conservé.
            code_brut = cfg.alias_dossiers.get(code_brut, code_brut)
            dossier = src.remap.get(code_brut, code_brut)
            # Si le code lu EST un code analytique remappé, il sert de centre ;
            # sinon le centre vient de la table dossier -> centre.
            centre = code_brut if code_brut in src.remap else cfg.analytique.get(dossier)
            if src.agreger:
                cumul[(dossier, centre)] += montant
            else:
                ajouter_ecriture_pair(par_dossier, dossier, src.journal, date, src.libelle,
                                      src.compte_credit, src.compte_debit, montant,
                                      centre, sans_centre, extourne, facteur=src.facteur,
                                      ventilation=src.ventilation.get(dossier),
                                      centres_connus=connus, centres_inconnus=centres_inconnus,
                                      fichier=src.fichier, numero_piece=numero_piece)

        for dossier, centre in sorted(cumul, key=lambda k: (k[0], k[1] or "")):
            ajouter_ecriture_pair(par_dossier, dossier, src.journal, date, src.libelle,
                                  src.compte_credit, src.compte_debit,
                                  round(cumul[(dossier, centre)], 2),
                                  centre, sans_centre, extourne, facteur=src.facteur,
                                  ventilation=src.ventilation.get(dossier),
                                  centres_connus=connus, centres_inconnus=centres_inconnus,
                                  fichier=src.fichier, numero_piece=numero_piece)
    return par_dossier, sans_centre


def generer_ecritures_paie(sources: List[SourcePaie], cfg: Configuration,
                           extourne: bool = False,
                           doublons: Optional[list] = None) -> Tuple[ParDossier, list, list]:
    """Traite les classeurs de paie détaillés par salarié.

    Les montants sont agrégés par centre de coût, le dossier est retrouvé via
    la table inverse centre -> dossier, et le centre de coût sert directement
    de centre analytique. Renvoie (par_dossier, centres_inconnus, en_attente),
    où centres_inconnus liste des couples (centre, fichier) pour la traçabilité.

    Si `doublons` (liste) est fourni, on y mémorise les clés (matricule, centre)
    présentes plus d'une fois (mêmes salarié+centre comptés deux fois, dans un
    ou plusieurs fichiers) sous la forme (matricule, centre, fichiers) — un même
    matricule sur des centres différents (salarié réparti) n'est PAS un doublon.
    """
    par_dossier: ParDossier = defaultdict(list)
    centres_inconnus: list = []
    en_attente: list = []
    suivi_matricules: Dict[Tuple[str, str], list] = defaultdict(list)
    cache_wb = {}

    for src in sources:
        if extourne and not src.contre_passation:
            continue
        chemin = os.path.join(cfg.dossier_entree, src.fichier)
        if chemin not in cache_wb:
            cache_wb[chemin] = load_workbook(chemin, data_only=True)
        ws = cache_wb[chemin][src.feuille]

        if src.entete_attendu:                       # contrôle de structure (bloquant)
            _verifier_entete(ws, src.entete_attendu,
                             src.ligne_entete or (src.ligne_debut - 1), src.fichier)

        c_centre = column_index_from_string(src.col_centre)
        c_matricule = column_index_from_string(src.col_matricule) if src.col_matricule else None
        date = src.contre_passation if extourne else src.date_ecriture
        numero_piece = src.numero_piece or cfg.numero_piece   # surcharge source > globale

        # Agrégation des montants par centre de coût, colonne par colonne (chaque
        # colonne lue une seule fois : plusieurs composantes peuvent partager la
        # même colonne, p. ex. brut + charges calculées au taux sur le même brut).
        cols_distinctes = {comp.col for comp in src.composantes}
        cumul: Dict[str, Dict[str, float]] = defaultdict(lambda: defaultdict(float))
        for r in range(src.ligne_debut, ws.max_row + 1):
            centre = ws.cell(r, c_centre).value
            if centre is None:
                continue
            centre = str(centre).strip()
            if not centre.isdigit():
                continue
            # Détection de doublons : on note (matricule, centre) -> fichier.
            if doublons is not None and c_matricule is not None:
                matricule = ws.cell(r, c_matricule).value
                if matricule is not None and str(matricule).strip():
                    suivi_matricules[(str(matricule).strip(), centre)].append(src.fichier)
            for col in cols_distinctes:
                v = lire_montant(ws.cell(r, column_index_from_string(col)).value)
                if v:
                    cumul[centre][col] += v

        for centre in sorted(cumul):
            dossier = cfg.centre_vers_dossier.get(centre)
            if dossier is None:
                centres_inconnus.append((centre, src.fichier))   # couple pour la traçabilité
                continue
            dossier = cfg.alias_dossiers.get(dossier, dossier)   # alias de dossier
            for comp in src.composantes:
                base = round(cumul[centre].get(comp.col, 0.0), 2)   # brut provisionné
                if comp.taux is not None:                           # charge = brut × taux
                    montant = float((Decimal(str(base)) * Decimal(str(comp.taux))
                                     ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
                else:
                    montant = base
                if abs(montant) < 0.005:
                    continue
                if not comp.complete:
                    en_attente.append(comp.libelle)
                    continue
                ajouter_ecriture_pair(par_dossier, dossier, src.journal, date, comp.libelle,
                                      comp.compte_credit, comp.compte_debit, montant,
                                      centre, None, extourne, numero_piece=numero_piece)

    if doublons is not None:                              # clés vues plus d'une fois
        for (matricule, centre), fichiers in sorted(suivi_matricules.items()):
            if len(fichiers) > 1:
                doublons.append((matricule, centre, tuple(sorted(fichiers))))
    return par_dossier, sorted(set(centres_inconnus)), sorted(set(en_attente))


def controler_equilibre(lignes: List[str]) -> Tuple[int, int]:
    """Renvoie (total débits, total crédits) en centimes — lignes M uniquement."""
    debit = credit = 0
    for l in lignes:
        if not l.startswith("M"):
            continue
        centimes = int(l[42:55])
        if l[41] == "D":
            debit += centimes
        else:
            credit += centimes
    return debit, credit


def formater_numero_piece(base: str, compteur: int) -> str:
    """Compose un n° de pièce « base + compteur » tenant sur 8 caractères.

    Le compteur est zéro-paddé sur 2 chiffres (3 au-delà de 99, etc.). En cas de
    dépassement des 8 caractères, c'est la **base** qui est rognée, jamais le
    compteur (la traçabilité du n° de run prime).
    """
    suffixe = str(compteur).zfill(2)
    base_max = max(0, 8 - len(suffixe))
    return (str(base)[:base_max] + suffixe)[:8]


def prochain_compteur(chemin: str) -> int:
    """Lit le dernier compteur, l'incrémente, le réécrit et renvoie la valeur du run.

    Fichier absent, vide ou corrompu : on repart de 1 (sans bloquer).
    """
    dernier = 0
    try:
        dernier = int(open(chemin, encoding="utf-8").read().strip())
    except (OSError, ValueError):
        dernier = 0                                  # absent / vide / corrompu
    courant = max(0, dernier) + 1
    dossier = os.path.dirname(chemin)
    if dossier:
        os.makedirs(dossier, exist_ok=True)
    with open(chemin, "w", encoding="utf-8") as fh:
        fh.write(str(courant))
    return courant


def archiver_entree(dossier_entree: str, dossier_archives: str,
                    horodatage: str) -> Optional[str]:
    """Archive les fichiers à la racine de `dossier_entree` dans un ZIP horodaté
    `entree_<horodatage>.zip` sous `dossier_archives` (créé si besoin).

    Renvoie le chemin du ZIP, ou None si `dossier_entree` est absent ou ne
    contient aucun fichier (pas de récursion dans les sous-dossiers).
    """
    if not os.path.isdir(dossier_entree):
        return None
    fichiers = [f for f in sorted(os.listdir(dossier_entree))
                if os.path.isfile(os.path.join(dossier_entree, f))]
    if not fichiers:
        return None
    os.makedirs(dossier_archives, exist_ok=True)
    chemin_zip = os.path.join(dossier_archives, f"entree_{horodatage}.zip")
    with zipfile.ZipFile(chemin_zip, "w", zipfile.ZIP_DEFLATED) as archive:
        for nom in fichiers:
            archive.write(os.path.join(dossier_entree, nom), arcname=nom)
    return chemin_zip


def nettoyer_sortie(dossier_sortie: str) -> list:
    """Supprime du dossier de sortie les seuls fichiers `*_ecriture_Quadra*.txt`.

    Évite qu'un fichier orphelin d'un run antérieur (dossier disparu, aliasé ou
    renommé) ne soit importé par erreur. Tout autre fichier est préservé. À
    appeler une fois au démarrage, avant la première écriture (les deux passes
    arrêté + contre-passation écrivent toutes deux des fichiers de ce motif).
    """
    if not os.path.isdir(dossier_sortie):
        return []
    supprimes = []
    for chemin in glob.glob(os.path.join(dossier_sortie, MOTIF_SORTIE)):
        os.remove(chemin)
        supprimes.append(os.path.basename(chemin))
    return supprimes


def ecrire_fichiers(par_dossier: ParDossier, dossier_sortie: str,
                    suffixe: str = "") -> Tuple[int, int, list]:
    """Écrit un fichier texte par dossier (cp1252, CRLF).

    Renvoie (total débits, total crédits, déséquilibres) en centimes.
    """
    os.makedirs(dossier_sortie, exist_ok=True)
    total_d = total_c = 0
    desequilibres = []
    for code in sorted(par_dossier):
        lignes = par_dossier[code]
        d, c = controler_equilibre(lignes)
        total_d += d
        total_c += c
        if d != c:
            desequilibres.append((code, d, c))
        nom = f"{code}_ecriture_Quadra{suffixe}.txt"
        with open(os.path.join(dossier_sortie, nom), "w",
                  encoding="cp1252", newline="") as fh:
            fh.write("\r\n".join(lignes) + "\r\n")
    return total_d, total_c, desequilibres
